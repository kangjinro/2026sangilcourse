import json, re, os
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, '..', 'data')
OUT_JS = os.path.join(SCRIPT_DIR, '..', 'js', 'data-embedded.js')

def clean(v):
    if v is None:
        return ''
    return str(v).strip()

def load_wb(path):
    return openpyxl.load_workbook(path, data_only=True)

def filled_aoa(ws):
    max_r, max_c = ws.max_row, ws.max_column
    aoa = [[ws.cell(row=r+1, column=c+1).value for c in range(max_c)] for r in range(max_r)]
    for m in ws.merged_cells.ranges:
        top = aoa[m.min_row-1][m.min_col-1]
        for r in range(m.min_row-1, m.max_row):
            for c in range(m.min_col-1, m.max_col):
                aoa[r][c] = top
    return aoa

# ---------------- 학과분류표 ----------------
def load_dept_class():
    wb = load_wb(f"{DATA_DIR}/departments_classification.xlsx")
    ws = wb[wb.sheetnames[0]]
    aoa = filled_aoa(ws)
    out = []
    for r in aoa[1:]:
        if r and clean(r[2] if len(r) > 2 else None):
            out.append({'major': clean(r[0]), 'middle': clean(r[1]), 'dept': clean(r[2])})
    return out

# ---------------- 학과정보 정리 ----------------
def load_dept_info():
    wb = load_wb(f"{DATA_DIR}/departments_info.xlsx")
    ws = wb[wb.sheetnames[0]]
    aoa = filled_aoa(ws)
    out = []
    for r in aoa[1:]:
        if r and clean(r[0] if len(r) > 0 else None):
            r = list(r) + [None]*10
            out.append({
                'name': clean(r[0]), 'intro': clean(r[1]), 'majorCourses': clean(r[2]),
                'recommendFor': clean(r[3]), 'similar': clean(r[4]), 'universities': clean(r[5]),
                'career': clean(r[6]), 'general': clean(r[7]), 'career_subj': clean(r[8]), 'fusion': clean(r[9]),
            })
    return out

# ---------------- 과목정보 정리 ----------------
def parse_daeyeok(text):
    if not text:
        return None
    m = re.match(r'대영역명\((.*?)\)\s*생각 열기 및 핵심 개념\((.*?)\)\s*주요 학습 활동예시\((.*)\)\s*$', text, re.S)
    if not m:
        return None
    return {
        'title': re.sub(r'\s+', ' ', m.group(1)).strip(),
        'hook': m.group(2).strip(),
        'activity': m.group(3).strip(),
    }

def load_subjects_info():
    wb = load_wb(f"{DATA_DIR}/subjects_info.xlsx")
    ws = wb[wb.sheetnames[0]]
    aoa = filled_aoa(ws)
    out = []
    for r in aoa[1:]:
        if not r or not clean(r[1] if len(r) > 1 else None):
            continue
        r = list(r) + [None]*17
        daeyeok_raw = [r[11], r[12], r[13], r[14], r[15], r[16]]
        daeyeok = [d for d in (parse_daeyeok(x) for x in daeyeok_raw) if d]
        out.append({
            'category': clean(r[0]), 'name': clean(r[1]), 'selectType': clean(r[2]),
            'absoluteScore': clean(r[3]), 'absoluteGrade': clean(r[4]), 'relativeGrade': clean(r[5]),
            'statDist': clean(r[6]), 'statAvg': clean(r[7]), 'statCount': clean(r[8]),
            'csat2029': clean(r[9]), 'introRaw': clean(r[10]), 'daeyeok': daeyeok,
        })
    return out

# ---------------- 2028 대입 권장과목(대교협) ----------------
def load_univ_recommend():
    wb = load_wb(f"{DATA_DIR}/university_recommend.xlsx")
    sheet_name = next((n for n in wb.sheetnames if '시트3' in n), wb.sheetnames[0])
    ws = wb[sheet_name]
    aoa = filled_aoa(ws)
    out = []
    for r in aoa[4:]:
        if not r or not clean(r[0] if len(r) > 0 else None):
            continue
        r = list(r) + [None]*8
        out.append({
            'region': clean(r[0]), 'area': clean(r[1]), 'university': clean(r[2]),
            'unitBroad': clean(r[3]), 'unitDept': clean(r[4]), 'coreSubjects': clean(r[5]),
            'recommendSubjects': clean(r[6]), 'note': clean(r[7]),
        })
    return out

# ---------------- 2026 상일고 교육과정 편성표 ----------------
def parse_select_type(raw):
    s = clean(raw)
    if not s:
        return {'kind': 'none'}
    if s == '지정':
        return {'kind': 'fixed'}
    m = re.search(r'그룹(\d)-택(\d)', s)
    if m:
        return {'kind': 'choice', 'group': int(m.group(1)), 'quota': int(m.group(2)), 'raw': s}
    return {'kind': 'other', 'raw': s}

def load_curriculum():
    wb = load_wb(f"{DATA_DIR}/curriculum_2026.xlsx")
    ws = wb[wb.sheetnames[0]]
    aoa = filled_aoa(ws)
    semester_defs = [
        {'key': '1-1', 'label': '1학년 1학기', 'startCol': 0},
        {'key': '1-2', 'label': '1학년 2학기', 'startCol': 6},
        {'key': '2-1', 'label': '2학년 1학기', 'startCol': 12},
        {'key': '2-2', 'label': '2학년 2학기', 'startCol': 18},
        {'key': '3-1', 'label': '3학년 1학기', 'startCol': 24},
        {'key': '3-2', 'label': '3학년 2학기', 'startCol': 30},
    ]
    by_semester = {d['key']: [] for d in semester_defs}
    for ri in range(4, len(aoa)):
        row = aoa[ri] or []
        for d in semester_defs:
            c = d['startCol']
            name = clean(row[c] if c < len(row) else None)
            if not name:
                continue
            select_type_raw = clean(row[c+1] if c+1 < len(row) else None)
            category = clean(row[c+2] if c+2 < len(row) else None)
            subject_type = clean(row[c+3] if c+3 < len(row) else None)
            credit_raw = row[c+4] if c+4 < len(row) else None
            try:
                credit = float(credit_raw)
                if credit == int(credit):
                    credit = int(credit)
            except (TypeError, ValueError):
                credit = 0
            by_semester[d['key']].append({
                'name': name, 'selectType': parse_select_type(select_type_raw),
                'selectTypeRaw': select_type_raw, 'category': category,
                'subjectType': subject_type, 'credit': credit,
            })
    return {'semesterDefs': semester_defs, 'bySemester': by_semester}

# ---------------- 2028 계열별 대표 모집단위별 반영과목 ----------------
def load_univ_recommend_series():
    wb = load_wb(f"{DATA_DIR}/univ_recommend_by_series.xlsx")
    sheet_name = next((n for n in wb.sheetnames if '반영과목' in n), wb.sheetnames[0])
    ws = wb[sheet_name]
    aoa = filled_aoa(ws)

    group_row = aoa[2] if len(aoa) > 2 else []
    sub_row = aoa[3] if len(aoa) > 3 else []
    total_cols = max(len(group_row), len(sub_row), 18)

    columns = []
    for c in range(2, min(total_cols, 18)):
        group = clean(group_row[c] if c < len(group_row) else None)
        sub = clean(sub_row[c] if c < len(sub_row) else None)
        if not group and not sub:
            continue
        columns.append({'idx': c, 'group': group or sub, 'sub': sub or group})

    def extract_university(row):
        for c in range(2, 17):
            v = clean(row[c] if c < len(row) else None)
            if v and v != '-':
                return v
        etc = clean(row[17] if len(row) > 17 else None)
        if etc and etc != '-':
            return etc.split('\n')[0]
        return '(대학명 미상)'

    records = []
    for r in aoa[4:]:
        if not r or not clean(r[1] if len(r) > 1 else None):
            continue
        records.append({
            'series': clean(r[0]),
            'unit': clean(r[1]),
            'university': extract_university(r),
            'cells': [clean(r[col['idx']] if col['idx'] < len(r) else None) or '-' for col in columns],
        })

    note = clean(aoa[1][0] if len(aoa) > 1 and len(aoa[1]) > 0 else None)
    return {'columns': columns, 'records': records, 'note': note}


if __name__ == '__main__':
    data = {
        'deptClass': load_dept_class(),
        'deptInfo': load_dept_info(),
        'subjectsInfo': load_subjects_info(),
        'univRecommend': load_univ_recommend(),
        'curriculum': load_curriculum(),
        'univRecommendSeries': load_univ_recommend_series(),
    }
    out_json = json.dumps(data, ensure_ascii=False)
    print("deptClass:", len(data['deptClass']))
    print("deptInfo:", len(data['deptInfo']))
    print("subjectsInfo:", len(data['subjectsInfo']))
    print("univRecommend:", len(data['univRecommend']))
    print("univRecommendSeries records:", len(data['univRecommendSeries']['records']))
    print("curriculum semesters:", list(data['curriculum']['bySemester'].keys()))
    print("JSON size (bytes):", len(out_json.encode('utf-8')))

    with open(OUT_JS, 'w', encoding='utf-8') as f:
        f.write('/* 자동 생성 파일 — 엑셀 데이터를 미리 변환해 내장한 데이터입니다.\n')
        f.write('   다시 생성하려면 이 스크립트(convert_to_json.py)를 실행하세요. 직접 수정하지 마세요. */\n')
        f.write('const EMBEDDED_APP_DATA = ' + out_json + ';\n')
    print("생성 완료:", OUT_JS)
